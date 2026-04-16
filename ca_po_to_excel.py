#!/usr/bin/env python3
"""
C&A Purchase Order PDF -> Excel Converter
Usage: ca_po_to_excel  <input.pdf>  [output.xlsx]
"""
import sys, re, os
from pathlib import Path
from datetime import date
from collections import defaultdict

try:
    import pdfplumber
except ImportError:
    print("ERROR: pdfplumber not installed.  pip install pdfplumber"); sys.exit(1)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed.  pip install openpyxl"); sys.exit(1)

# ── helpers ───────────────────────────────────────────────────────────────────
def _val(txt, pat, default="", flags=0):
    m = re.search(pat, txt, flags)
    return m.group(1).strip() if m else default

def _int(s):
    try: return int(str(s).replace(",","").strip())
    except: return 0

MONTH_MAP = {"jan":1,"feb":2,"mar":3,"apr":4,"may":5,"jun":6,
             "jul":7,"aug":8,"sep":9,"oct":10,"nov":11,"dec":12}

def _parse_date(s):
    m = re.match(r"(\d+)\.(\w+)\.(\d{4})", s.strip())
    if not m: return None
    return date(int(m.group(3)), MONTH_MAP.get(m.group(2).lower(),1), int(m.group(1)))

def _wbl(page):
    d = defaultdict(list)
    for w in page.extract_words(): d[round(w["top"])].append(w)
    return dict(d)

COUNTRY_NAMES = {
    "NL":"NETHERLAND","D":"GERMANY","B":"BELGIUM","F":"FRANCE",
    "CH":"SWITZERLAND","E":"SPAIN","A":"AUSTRIA","SK":"SLOVAKIA",
    "OL":"OL (INTERNET COMPANY)",
}
COUNTRY_ORDER = ["NL","D","B","F","CH","E","A","SK","OL"]
PACK_LETTERS  = list("ABCDEFGHIJKLMN")
COUNTRIES_SET = set(COUNTRY_NAMES.keys())

# ── page 1: header ────────────────────────────────────────────────────────────
def parse_header(txt):
    po_log    = _val(txt, r"(\d{5}-\d{3}-\d{2}-\d{3})")
    po_number = _val(txt, r"PONo\.:(\d+)")
    raw       = _val(txt, r"StyleName:(.+?)(?:SupplierStyle|Season)", flags=re.DOTALL)
    style     = re.sub(r"\s+", "", raw)
    date_m    = re.search(r"(\d+\.\w+\.\d{4})", txt)
    ship_date = _parse_date(date_m.group(1)) if date_m else None
    return dict(po_log=po_log, po_number=po_number,
                style_name=style, ship_date=ship_date)

# ── page 5: per-country TSS counts ───────────────────────────────────────────
def parse_tss_counts(page):
    wbl = _wbl(page)
    tss_y0 = tss_y1 = None
    for top, words in sorted(wbl.items()):
        line = " ".join(w["text"] for w in words)
        if "PacksTSS"  in line and tss_y0 is None: tss_y0 = top
        if "PacksTNSS" in line and tss_y0:          tss_y1 = top; break
    counts = {}
    if not (tss_y0 and tss_y1): return counts
    for top, words in sorted(wbl.items()):
        if not (tss_y0 < top < tss_y1): continue
        sw = sorted(words, key=lambda w: w["x0"])
        if not sw or sw[0]["text"] not in COUNTRIES_SET: continue
        cc   = sw[0]["text"]
        nums = [w["text"] for w in sw[1:] if re.match(r"[\d,]+$", w["text"])]
        if nums: counts[cc] = _int(nums[0])
    return counts

# ── pages 6-7: pack-country map + size data ───────────────────────────────────
def parse_pack_data(pages):
    """
    Returns pack_countries {letter:[cc,...]}, tss_ratios {letter:sizes},
    tnss_sizes {letter:sizes}, color_name.
    
    Key insight: country annotation rows like "(NL) (D)" appear 14 y-units
    ABOVE the ratio line, which is 12 y-units above the size data line.
    Only look BELOW each annotation row within a narrow ±50 y window.
    """
    pack_countries = {}
    tss_ratios     = {}
    tnss_sizes     = {}
    color_name     = "DUSTY OLIVE"
    pair_idx       = 0

    for page in pages:
        wbl  = _wbl(page)
        rows = sorted(wbl.items())

        # Colour
        txt = page.extract_text() or ""
        cm = re.search(r"\d{2}\s+\d+,([^\d\n]+)\s+[\d,]+\s+[\d,]+\s+[\d,]+\s+[\d,]+\s+[\d,]+\s+[\d,]+", txt)
        if cm: color_name = cm.group(1).strip().upper()

        for top, words in rows:
            line = " ".join(w["text"] for w in sorted(words, key=lambda w: w["x0"]))
            cg = re.findall(r"\(([A-Z,]+)\)", line)
            if len(cg) < 2: continue
            if pair_idx + 1 >= len(PACK_LETTERS): continue

            left_ccs  = [c.strip() for c in cg[0].split(",")]
            right_ccs = [c.strip() for c in cg[1].split(",")]
            left_ltr  = PACK_LETTERS[pair_idx]
            right_ltr = PACK_LETTERS[pair_idx+1]
            pack_countries[left_ltr]  = left_ccs
            pack_countries[right_ltr] = right_ccs

            # Find ratio & size lines strictly BELOW this row (top < other_top < top+50)
            ratio_seen = {}   # letter → n_packs
            for other_top, other_words in rows:
                if other_top <= top or other_top > top + 55: continue
                ow = sorted(other_words, key=lambda w: w["x0"])

                # Ratio line
                for w in ow:
                    rm = re.search(r"(\d[\d,]*)([A-N])Packsof(\d[\d,]*)Pieces=(\d[\d,]*)Pieces",
                                   w["text"])
                    if rm:
                        ltr  = rm.group(2)
                        n    = _int(rm.group(1))
                        ratio_seen[ltr] = n

                # Size line: "01" article code at far left
                if ow and ow[0]["text"] == "01":
                    # Skip x<100 (article code "01" and colour text)
                    left_nums  = [_int(w["text"]) for w in ow
                                  if w["x0"] >= 100 and w["x0"] < 400
                                  and re.match(r"[\d,]+$", w["text"])]
                    right_nums = [_int(w["text"]) for w in ow
                                  if w["x0"] >= 450
                                  and re.match(r"[\d,]+$", w["text"])]

                    def to_sz(nums, ltr):
                        # nums = [total, xs, s, m, l, xl, xxl] (7 values)
                        # or     [total, xs, s, m, xl, xxl]    (6 values, missing l=0)
                        if len(nums) >= 7:
                            return dict(xs=nums[1],s=nums[2],m=nums[3],
                                        l=nums[4],xl=nums[5],xxl=nums[6])
                        if len(nums) == 6:
                            # Insert 0 for missing size (check sum)
                            # Try inserting 0 at L position
                            s = dict(xs=nums[1],s=nums[2],m=nums[3],l=0,xl=nums[4],xxl=nums[5])
                            return s
                        return None

                    for ltr_pair, nums in [(left_ltr, left_nums), (right_ltr, right_nums)]:
                        sz = to_sz(nums, ltr_pair)
                        if sz is None: continue
                        n = ratio_seen.get(ltr_pair, 0)
                        if n > 1:
                            tss_ratios[ltr_pair] = sz
                        else:
                            tnss_sizes[ltr_pair] = sz

            pair_idx += 2

    return pack_countries, tss_ratios, tnss_sizes, color_name

# ── build pack_counts ─────────────────────────────────────────────────────────
def build_pack_counts(pack_countries, tss_counts, tss_ratios, tnss_sizes):
    counts = {}
    for letter, ccs in pack_countries.items():
        if letter in tss_ratios:
            for cc in ccs:
                n = tss_counts.get(cc, 0)
                if n: counts.setdefault(cc, {})[letter] = n
        elif letter in tnss_sizes:
            for cc in ccs:
                counts.setdefault(cc, {})[letter] = 1
    return counts

# ── Excel builder ─────────────────────────────────────────────────────────────
_thin      = Side(style="thin", color="000000")
ALL_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
HDR_FONT   = Font(name="Tahoma", size=8, bold=True)
DATA_FONT  = Font(name="Tahoma", size=8, bold=False)
HDR_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left",   vertical="center")
RIGHT_ALIGN= Alignment(horizontal="right",  vertical="center")
CTR_ALIGN  = Alignment(horizontal="center", vertical="center")

HEADERS    = [
    "S No","O.Style","O.Ponum","O.Ponum Log.","Ship Date",
    "Pack Type","Pack Ratio","D.Country","Country Code",
    "O.Combo","O.Sizes","O.Qty",
    "XS","S","M","L","XL","XXL",
    "","","","","","","","","", "Total Qty", ""
]
COL_WIDTHS = [
    6.0, 20.9, 17.0, 17.3, 9.9, 9.7, 8.7, 19.3, 8.7, 36.7, 9.9, 9.7,
] + [8.7]*15 + [8.7, 2.1]

def _wcell(ws, row, col, val):
    cell = ws.cell(row=row, column=col, value=val)
    cell.border = ALL_BORDER; cell.font = DATA_FONT
    if isinstance(val, date):
        cell.number_format = "DD-MMM-YY"; cell.alignment = CTR_ALIGN
    elif isinstance(val, (int,float)) and col > 1:
        cell.number_format = "#,##0"; cell.alignment = RIGHT_ALIGN
    else:
        cell.alignment = LEFT_ALIGN

def build_excel(hdr, pack_counts, tss_ratios, tnss_sizes, color_name, out_path):
    wb = Workbook(); ws = wb.active; ws.title = "Sheet1"
    ws.row_dimensions[1].height = 20
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HDR_FONT; cell.alignment = HDR_ALIGN; cell.border = ALL_BORDER
        ws.column_dimensions[get_column_letter(c)].width = COL_WIDTHS[c-1]

    sno = 1; row = 2

    def add_row(cc, letter, sizes, total):
        nonlocal sno, row
        vals = [
            sno, hdr["style_name"], f"{hdr['po_number']}-{cc}",
            hdr["po_log"]+"-001", hdr["ship_date"],
            f"{letter}-PACK", 0,
            COUNTRY_NAMES.get(cc, cc), cc, color_name, "", "",
            sizes["xs"],sizes["s"],sizes["m"],sizes["l"],sizes["xl"],sizes["xxl"],
            0,0,0,0,0,0,0,0,0, total, ""
        ]
        ws.row_dimensions[row].height = 15
        for c2, v in enumerate(vals, 1): _wcell(ws, row, c2, v)
        sno += 1; row += 1

    for cc in COUNTRY_ORDER:
        if cc not in pack_counts: continue
        for letter in PACK_LETTERS:
            count = pack_counts[cc].get(letter, 0)
            if not count: continue
            if letter in tss_ratios:
                r = tss_ratios[letter]
                sizes = {k: r[k]*count for k in ("xs","s","m","l","xl","xxl")}
                add_row(cc, letter, sizes, sum(sizes.values()))

    for cc in COUNTRY_ORDER:
        if cc not in pack_counts: continue
        for letter in PACK_LETTERS:
            count = pack_counts[cc].get(letter, 0)
            if not count: continue
            if letter in tnss_sizes:
                add_row(cc, letter, tnss_sizes[letter],
                        sum(tnss_sizes[letter].values()))

    wb.save(out_path)
    return sno - 1

# ── orchestrator ──────────────────────────────────────────────────────────────
def pdf_to_excel(pdf_path, xlsx_path):
    with pdfplumber.open(pdf_path) as pdf:
        p = pdf.pages
        hdr        = parse_header(p[0].extract_text() or "")
        tss_counts = parse_tss_counts(p[4])
        pack_countries, tss_ratios, tnss_sizes, color = parse_pack_data(p[5:7])
    pack_counts = build_pack_counts(pack_countries, tss_counts, tss_ratios, tnss_sizes)
    rows = build_excel(hdr, pack_counts, tss_ratios, tnss_sizes, color, xlsx_path)
    return rows, hdr

def main():
    if len(sys.argv) < 2:
        print("C&A PO PDF → Excel Converter")
        print("Usage: ca_po_to_excel <input.pdf> [output.xlsx]"); sys.exit(1)
    pdf_path = sys.argv[1]
    if not os.path.isfile(pdf_path):
        print(f"ERROR: file not found: {pdf_path}"); sys.exit(1)
    xlsx_path = sys.argv[2] if len(sys.argv) >= 3 \
                else str(Path(pdf_path).with_suffix(".xlsx"))
    print(f"Reading : {pdf_path}")
    try:
        rows, hdr = pdf_to_excel(pdf_path, xlsx_path)
    except Exception as e:
        import traceback; traceback.print_exc(); sys.exit(1)
    print(f"Saved   : {xlsx_path}")
    print(f"PO #    : {hdr['po_number']}   Style: {hdr['style_name'][:50]}")
    print(f"Rows    : {rows}")

if __name__ == "__main__":
    main()
